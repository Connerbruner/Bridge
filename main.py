import os
import discord
import threading
import time
import openpyxl
from datetime import date
from slack_sdk import WebClient
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from discord.ext import tasks

slack = WebClient(token=os.environ.get("slack"))
app = App(token=os.environ.get("slack"),signing_secret="")

intents = discord.Intents.default()
intents.message_content = True

client = discord.Client(intents=intents)

def commands(msg,bool,channel):
    if "!msg" in msg:
        message(msg[4:],not bool,channel)
        return False
    elif "!read" in msg:
        with open(channel) as file:
            for line in file:
                if msg[5:] in line:
                    message(line,bool,channel)
        return False  
    elif "!wakeup" in msg:
        if bool:
            i=25
            while i>0:
                message("WAKE UP",True,msg[10:len(str)-1])
                time.sleep(0.5)
                i-=1
    elif "!help" in msg:
        message("!msg messages other app\n!find finds item in sorting system\n!read (slack only) searchs channel archive for the word you are looking for\n!wakeup (slack only) sends user 25 DMS",bool,channel)
        return False
    elif "!find" in msg:
        wrkbk = openpyxl.load_workbook("sorting.xlsx")
    
        sh = wrkbk.active

        for i in range(2, sh.max_row+1):      
            for j in range(1, sh.max_column+1):
                if msg[6:] in str(sh.cell(row=i, column=j).value):
                    message(sh.cell(row=i, column=j).value+" can be found in "+sh.cell(row=i, column=1).value,bool,channel)
        return False


    elif "<!everyone>" in msg and bool:
        message("New announcement in slack",not bool,channel)
    
    return True
        
def message(str,bool,channel):
    if bool:
        slack.chat_postMessage(channel=channel,text=str)
    else:
        channel = client.get_channel(912762550679142442)
        client.loop.create_task(channel.send(str))
        
@client.event
async def on_ready():
    printer.start()
    
@client.event
async def on_message(message):
    if message.author == client.user:
        return
    commands(message.content,False,"#bridge-channel")
        
        
@app.event('message')
def print_message(event):
    if commands(event['text'],True,event['channel']):
        file = open(event['channel'],"a")
        file.write("["+str(date.today())+"]<@"+event['user']+">: "+event['text']+"\n")
        file.close()
    
    
@tasks.loop(seconds=1)
async def printer():
        return

def discord():
    client.run(os.environ["discord"])

t = threading.Thread(target=discord, args=())
t.start()
SocketModeHandler(app, os.environ["app"]).start()
